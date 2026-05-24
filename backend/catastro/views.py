from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth import authenticate, login
from rest_framework_simplejwt.tokens import RefreshToken
from openpyxl.worksheet.datavalidation import DataValidation
from .utils import valida_rut_chileno, valida_nombre, send_email_async
from django.contrib.auth import get_user_model
from django.urls import reverse
from .models import Condominio, Unidad, Residente, Documento, Notificacion
from .serializers import CondominioSerializer, UnidadSerializer, ResidenteSerializer, DocumentoSerializer, NotificacionSerializer

class RoleBasedAccess(permissions.BasePermission):
    """
    Permiso custom para proteger los datos médicos:
    - Admin y Conserje pueden ver el listado (para evacuaciones).
    - Un Residente solo interactúa con los de su misma unidad.
    """
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated)

class CondominioViewSet(viewsets.ModelViewSet):
    queryset = Condominio.objects.all()
    serializer_class = CondominioSerializer
    permission_classes = [permissions.IsAuthenticated]

class UnidadViewSet(viewsets.ModelViewSet):
    queryset = Unidad.objects.all()
    serializer_class = UnidadSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['post'], url_path='carga-masiva')
    def carga_masiva(self, request):
        """Carga masiva de unidades desde archivo Excel."""
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'detail': 'Debe adjuntar un archivo Excel.'}, status=status.HTTP_400_BAD_REQUEST)
        
        import openpyxl
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception as e:
            return Response({'detail': f'Error al leer el archivo: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Leer encabezados
        headers = [cell.value.strip().lower() if cell.value else '' for cell in ws[1]]
        required = ['numero_depto']
        if 'numero_depto' not in headers:
            return Response({'detail': 'Falta la columna requerida: numero_depto'}, status=status.HTTP_400_BAD_REQUEST)
        
        condominio = Condominio.objects.first()
        if not condominio:
            return Response({'detail': 'No hay condominio configurado.'}, status=status.HTTP_400_BAD_REQUEST)
            
        creados = 0
        errores = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, [str(v).strip() if v else '' for v in row]))
            
            if not data.get('numero_depto'):
                errores.append(f'Fila {row_idx}: numero_depto es obligatorio')
                continue
                
            try:
                unidad, created = Unidad.objects.get_or_create(
                    condominio=condominio,
                    torre=data.get('torre', ''),
                    numero_depto=data['numero_depto']
                )
                if created:
                    creados += 1
            except Exception as e:
                errores.append(f'Fila {row_idx}: {str(e)}')
                
        return Response({
            'detail': f'Carga completada. {creados} unidades creadas.',
            'creados': creados,
            'errores': errores
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['delete'], url_path='eliminar-todas')
    def eliminar_todas(self, request):
        """Elimina todas las unidades y sus dependencias."""
        count, _ = Unidad.objects.all().delete()
        return Response({'detail': f'Se han eliminado {count} unidades.'}, status=status.HTTP_200_OK)

class ResidenteViewSet(viewsets.ModelViewSet):
    queryset = Residente.objects.all()
    serializer_class = ResidenteSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if getattr(user, 'rol', None) == 'RESIDENTE':
            # Buscar la unidad del residente logueado y devolver todos los de esa unidad
            mi_residente = Residente.objects.filter(correo=user.email).first()
            if mi_residente and mi_residente.unidad_id:
                return Residente.objects.filter(unidad=mi_residente.unidad)
            return Residente.objects.filter(correo=user.email)
        return super().get_queryset()

    @action(detail=True, methods=['post'], url_path='enviar-invitacion')
    def enviar_invitacion(self, request, pk=None):
        """Genera una invitación por correo para que el residente cree su contraseña."""
        try:
            residente = self.get_object()
        except Exception as e:
            return Response({'detail': 'Residente no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        User = get_user_model()
        # Crear/obtener usuario asociado al correo del residente
        user, created = User.objects.get_or_create(username=residente.correo, defaults={
            'email': residente.correo,
            'rol': User.Roles.RESIDENTE,
            'is_active': True,
        })
        # Generar token y uid
        token_generator = PasswordResetTokenGenerator()
        token = token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        # Construir enlace (asumimos que frontend corre en localhost:5173)
        link = f"http://localhost:5173/establecer-clave/{uid}/{token}"
        # Enviar email
        subject = "Invitación de acceso a CondoConnect"
        message = f"Hola {residente.nombre},\n\nSe ha creado una cuenta para ti en el portal.\nHaz clic en el siguiente enlace para establecer tu contraseña y activar tu acceso:\n{link}\n\nSi no solicitaste esto, ignora el mensaje."
        send_email_async(subject=subject, message=message, recipient_list=[residente.correo])
        return Response({'detail': 'Invitación enviada correctamente.'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='restablecer-clave')
    def restablecer_clave(self, request, pk=None):
        """Envía email de restablecimiento de contraseña al residente."""
        residente = self.get_object()
        if not residente.correo:
            return Response({'detail': 'El residente no tiene correo registrado.'}, status=status.HTTP_400_BAD_REQUEST)
        User = get_user_model()
        try:
            user = User.objects.get(email=residente.correo)
        except User.DoesNotExist:
            return Response({'detail': 'No se encontró una cuenta para este correo.'}, status=status.HTTP_404_NOT_FOUND)
        token_generator = PasswordResetTokenGenerator()
        token = token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        link = f"http://localhost:5173/establecer-clave/{uid}/{token}"
        subject = "Restablecer contraseña - CondoConnect"
        message = f"Hola,\n\nHaz clic en el siguiente enlace para restablecer tu contraseña:\n{link}\n\nSi no solicitaste esto, ignora el mensaje."
        send_email_async(subject=subject, message=message, recipient_list=[residente.correo])
        return Response({'detail': 'Correo de restablecimiento enviado.'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['delete'], url_path='eliminar-cuenta')
    def eliminar_cuenta(self, request, pk=None):
        """Elimina la cuenta de usuario (CustomUser) asociada al residente, sin borrar la ficha de residente."""
        residente = self.get_object()
        if not residente.correo:
            return Response({'detail': 'El residente no tiene correo registrado.'}, status=status.HTTP_400_BAD_REQUEST)
        User = get_user_model()
        try:
            user = User.objects.get(email=residente.correo)
            user.delete()
            return Response({'detail': 'Cuenta de usuario eliminada correctamente.'}, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'detail': 'No se encontró una cuenta de usuario para este residente.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'], url_path='enviar-soporte')
    def enviar_soporte(self, request):
        """Envía un correo de soporte al administrador del condominio."""
        mensaje = request.data.get('mensaje')
        if not mensaje:
            return Response({'detail': 'El mensaje no puede estar vacío.'}, status=status.HTTP_400_BAD_REQUEST)
        
        user = request.user
        residente = Residente.objects.filter(correo=user.email).first()
        nombre_remitente = f"{residente.nombre} {residente.apellidos or ''}".strip() if residente else user.email
        correo_remitente = user.email

        condominio = Condominio.objects.first()
        if not condominio or not condominio.email_administracion:
            return Response({'detail': 'No hay un correo de administración configurado.'}, status=status.HTTP_400_BAD_REQUEST)
            
        subject = f"Nuevo mensaje de soporte - {nombre_remitente}"
        cuerpo = f"Has recibido un nuevo mensaje de soporte desde la plataforma.\n\nRemitente: {nombre_remitente}\nCorreo: {correo_remitente}\n\nMensaje:\n{mensaje}"
        
        send_email_async(subject=subject, message=cuerpo, recipient_list=[condominio.email_administracion])
        return Response({'detail': 'Mensaje enviado correctamente.'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='confirmar-datos')
    def confirmar_datos(self, request, pk=None):
        """Confirma que los datos del residente están actualizados (reinicia contador semestral)."""
        residente = self.get_object()
        residente.save()
        return Response({'detail': 'Datos confirmados correctamente.'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='carga-masiva')
    def carga_masiva(self, request):
        """Carga masiva de residentes desde archivo Excel."""
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'detail': 'Debe adjuntar un archivo Excel.'}, status=status.HTTP_400_BAD_REQUEST)
        
        import openpyxl
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception as e:
            return Response({'detail': f'Error al leer el archivo: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Leer encabezados
        headers = [cell.value.strip().lower() if cell.value else '' for cell in ws[1]]
        required = ['numero_depto', 'nombres']
        for col in required:
            if col not in headers:
                return Response({'detail': f'Falta la columna requerida: {col}'}, status=status.HTTP_400_BAD_REQUEST)
        
        User = get_user_model()
        token_generator = PasswordResetTokenGenerator()
        creados = 0
        errores = []
        invitaciones = 0
        condominio = Condominio.objects.first()
        
        if not condominio:
            return Response({'detail': 'No hay condominio configurado.'}, status=status.HTTP_400_BAD_REQUEST)
        
        VALID_RELACIONES = ['JEFE_HOGAR', 'CONYUGE', 'ARRENDATARIO', 'FAMILIAR_MENOR', 'FAMILIAR_ADULTO', 'FAMILIAR_MAYOR', 'OTRO']

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Check if row is completely empty
            if all(v is None or str(v).strip() == '' for v in row):
                continue
                
            data = dict(zip(headers, [str(v).strip() if v is not None else '' for v in row]))
            
            nombres = data.get('nombres', '')
            numero_depto = data.get('numero_depto', '')
            
            if no_nombres or not numero_depto:
                errores.append(f'Fila {row_idx}: nombres y numero_depto son obligatorios.')
                continue
                
            if not valida_nombre(nombres):
                errores.append(f'Fila {row_idx}: Nombres no pueden contener números o caracteres especiales.')
                continue
                
            apellidos = data.get('apellidos', '')
            if apellidos and not valida_nombre(apellidos):
                errores.append(f'Fila {row_idx}: Apellidos no pueden contener números o caracteres especiales.')
                continue
                
            rut_dni = data.get('rut_dni', '')
            if rut_dni and not valida_rut_chileno(rut_dni):
                errores.append(f'Fila {row_idx}: RUT {rut_dni} no es válido.')
                continue
                
            relacion = data.get('relacion_hogar', 'JEFE_HOGAR').upper()
            if relacion and relacion not in VALID_RELACIONES:
                errores.append(f"Fila {row_idx}: Relación '{relacion}' no es válida.")
                continue
                
            movilidad = data.get('movilidad_reducida', '').lower() == 'sí' or data.get('movilidad_reducida', '').lower() == 'si'
            medica = data.get('condicion_medica', '').lower() == 'sí' or data.get('condicion_medica', '').lower() == 'si'
            
            # Format fecha_nacimiento
            fecha_nacimiento = data.get('fecha_nacimiento')
            if fecha_nacimiento:
                try:
                    # In case it's a datetime object string representation or yyyy-mm-dd
                    if ' ' in fecha_nacimiento:
                        fecha_nacimiento = fecha_nacimiento.split(' ')[0]
                except Exception:
                    fecha_nacimiento = None
            else:
                fecha_nacimiento = None
            
            try:
                # Buscar unidad, considerando que torre puede ser None o ''
                torre_str = data.get('torre', '')
                if torre_str:
                    unidad = Unidad.objects.filter(condominio=condominio, torre=torre_str, numero_depto=numero_depto).first()
                else:
                    unidad = Unidad.objects.filter(condominio=condominio, numero_depto=numero_depto, torre__in=[None, '']).first()
                
                if not unidad:
                    errores.append(f"Fila {row_idx}: La unidad Depto {numero_depto} (Torre {torre_str or 'N/A'}) no existe. Debe crear la unidad primero.")
                    continue
                
                correo = data.get('correo', '')
                if correo:
                    correo = correo.lower()
                
                # Crear residente
                residente = Residente.objects.create(
                    unidad=unidad,
                    nombre=nombres,
                    apellidos=data.get('apellidos', ''),
                    rut_dni=data.get('rut_dni', ''),
                    fecha_nacimiento=fecha_nacimiento or None,
                    nacionalidad=data.get('nacionalidad', ''),
                    idioma_principal=data.get('idioma', ''),
                    correo=correo or None,
                    telefono=data.get('telefono', ''),
                    movilidad_reducida=movilidad,
                    condicion_medica=medica,
                    relacion_jefe_hogar=relacion or 'JEFE_HOGAR',
                    contacto_emergencia_nombre=data.get('contacto_emergencia_nombre', ''),
                    contacto_emergencia_telefono=data.get('contacto_emergencia_telefono', ''),
                    contacto_emergencia_correo=data.get('contacto_emergencia_correo', '') or None
                )
                creados += 1
                
                # Crear usuario y enviar invitación si tiene correo
                if correo:
                    user, created = User.objects.get_or_create(
                        username=correo,
                        defaults={
                            'email': correo,
                            'rol': User.Roles.RESIDENTE,
                            'is_active': True,
                        }
                    )
                    if created:
                        token = token_generator.make_token(user)
                        uid = urlsafe_base64_encode(force_bytes(user.pk))
                        link = f"http://localhost:5173/establecer-clave/{uid}/{token}"
                        subject = "Invitación de acceso a CondoConnect"
                        message = f"Hola {nombres},\n\nSe ha creado una cuenta para ti en el portal.\nHaz clic en el siguiente enlace para establecer tu contraseña:\n{link}\n\nSi no solicitaste esto, ignora el mensaje."
                        send_email_async(subject=subject, message=message, recipient_list=[correo])
                        invitaciones += 1
                    
            except Exception as e:
                errores.append(f'Fila {row_idx}: Error interno al procesar ({str(e)})')
        
        return Response({
            'detail': f'Carga completada. {creados} residentes creados, {invitaciones} invitaciones enviadas.',
            'creados': creados,
            'invitaciones': invitaciones,
            'errores': errores
        }, status=status.HTTP_200_OK)

class EstablecerClaveView(viewsets.ViewSet):
    """Endpoint que valida token y permite crear la contraseña del usuario."""
    def create(self, request):
        uidb64 = request.data.get('uid')
        token = request.data.get('token')
        password = request.data.get('password')
        if not all([uidb64, token, password]):
            return Response({'detail': 'Datos incompletos.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            uid = force_str(urlsafe_base64_decode(uidb64))
            User = get_user_model()
            user = User.objects.get(pk=uid)
        except Exception:
            return Response({'detail': 'Usuario no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        token_generator = PasswordResetTokenGenerator()
        if not token_generator.check_token(user, token):
            return Response({'detail': 'Token inválido o expirado.'}, status=status.HTTP_400_BAD_REQUEST)
        user.set_password(password)
        user.save()
        return Response({'detail': 'Contraseña establecida correctamente.'}, status=status.HTTP_200_OK)


class RecuperarClaveView(viewsets.ViewSet):
    """Endpoint para solicitar recuperación de contraseña por correo."""
    def create(self, request):
        email = request.data.get('email', '').strip()
        if not email:
            return Response({'detail': 'Debe ingresar un correo electrónico.'}, status=status.HTTP_400_BAD_REQUEST)
        User = get_user_model()
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            # Por seguridad, no revelamos si el correo existe o no
            return Response({'detail': 'Si el correo está registrado, recibirá un enlace para restablecer su contraseña.'}, status=status.HTTP_200_OK)
        token_generator = PasswordResetTokenGenerator()
        token = token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        link = f"http://localhost:5173/establecer-clave/{uid}/{token}"
        subject = "Recuperación de contraseña - CondoConnect"
        message = (
            f"Hola {user.get_full_name() or user.username},\n\n"
            f"Recibimos una solicitud para restablecer tu contraseña.\n"
            f"Haz clic en el siguiente enlace para crear una nueva contraseña:\n{link}\n\n"
            f"Si no solicitaste esto, puedes ignorar este mensaje.\n"
            f"El enlace expirará en 24 horas."
        )
        send_email_async(subject=subject, message=message, recipient_list=[email])
        return Response({'detail': 'Si el correo está registrado, recibirá un enlace para restablecer su contraseña.'}, status=status.HTTP_200_OK)

class DocumentoViewSet(viewsets.ModelViewSet):
    queryset = Documento.objects.all()
    serializer_class = DocumentoSerializer
    permission_classes = [permissions.IsAuthenticated]

from .utils import check_vencimientos_sistema

class NotificacionViewSet(viewsets.ModelViewSet):
    queryset = Notificacion.objects.all()
    serializer_class = NotificacionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        # Corremos el check automático cada vez que se consultan las notificaciones
        # (En un entorno real de alta carga esto iría en Celery o Cron, aquí on-the-fly es perfecto)
        try:
            check_vencimientos_sistema()
        except Exception as e:
            print(f"Error checking vencimientos: {e}")

        user = self.request.user
        qs = Notificacion.objects.all()
        # Si el usuario tiene rol Residente, solo ve las notificaciones de su unidad o globales
        if getattr(user, 'rol', None) == 'RESIDENTE':
            # Tenemos que saber de qué unidad es, busquemos su registro Residente
            # Asumimos que el correo del User coincide con el del Residente
            residentes_mios = Residente.objects.filter(correo=user.email)
            unidades_ids = residentes_mios.values_list('unidad_id', flat=True)
            # Retorna notificaciones de esas unidades o globales (unidad=None)
            from django.db.models import Q
            qs = qs.filter(Q(unidad__in=unidades_ids) | Q(unidad__isnull=True)).exclude(titulo="Residente eliminado")
        return qs
